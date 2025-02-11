from openpyxl import load_workbook

client_list = 'suricatos.xlsx'

wb = load_workbook(client_list)
ws = wb.active

def build_code():
    return 1000+ws.max_row+1

def append_client(client_name,client_mail,client_code):
    new_data = [[client_name,client_mail,client_code]]
    for row in new_data:
        ws.append(row)

    wb.save(client_list)
